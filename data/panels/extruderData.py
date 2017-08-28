#!/usr/bin/python

'''
Program: Interfacial Consultant's Systems and Management - ICSM
Programmer: Talib M. Khan
Date Created: 08/07/2017
Last Updated: 08/24/2017
Version: 1.0.0
Description:
    The following python file contains the model code functions for the "Extruder" panel for the ICSM program
'''

'''
Imported files/libraries
'''
import os
import sys
import json
from types import ModuleType
import openpyxl
from openpyxl.styles import colors
from openpyxl.styles import Border, Side
from add import feederData

'''
Global variables
'''
ERROR = "AN ERROR HAS OCCURRED"

'''
The following class is the model for the "Extruder" panel
'''
class ExtruderD:

  '''
  The following function is the initial instance creation function for the "ExtruderD" class
  '''
  def __init__(self, config, socket):

    # Call set functions and if the return is False, then return the message for the error
    doesWork, message = self.setConfig(config)
    if not doesWork:
      print message
      sys.exit()

    # Set initial values for this instance of the "ExtruderD" class
    self.graphics = None
    self.socket = socket
    self.workflowFileName = None
    self.extruderVariable = None
    self.dieVariable = None
    self.preDieVariable = None
    self.screenVariable = None
    self.portVariables = []
    self.sideVariables = []
    self.feeders = {}
    self.strandCoolingFrameVariable = None
    self.separatorVariable = None
    self.fansVariable = None
    self.airKnivesVariable = None
    self.numAirKnivesVariable = None
    self.beltMisters = []
    self.sluiceMisters = []
    self.pelletizierVariable = None
    self.classifiedVariable = None
    self.dataShiftVariable = None
    self.numDataVariable = None
    self.fromVariable = None
    self.toVariable = None

  '''
  The following function returns the configuration file for this instance of the "ExtruderD" class
  '''
  def getConfig(self):
    return self.config

  '''
  The following function sets the configuration file for this instance of the class. If the inputted config file does
  not meet the requirements then the function returns a "False" boolean value and an error message
  '''
  def setConfig(self, config):
    if isinstance(config, ModuleType):
      if hasattr(config, "confirm"):
        if config.confirm("Extruder"):
          self.config = config
        else:
          return False, "%s:\nconfig file for ExtruderD is not configExtruder" % ERROR
      else:
        return False, "%s:\nconfig file is not a designated config file for this program" % ERROR
    else:
      return False, "%s:\ninputted file is not a Module" % ERROR
    return True, ""

  '''
  The following function returns the view for this instance of the "Extruder" panel
  '''
  def getGraphics(self):
    return self.graphics

  '''
  The following function sets the view for this instance of the "Extruder" panel. If the input does not meet the
  requirements then the function returns a "False" boolean value and an error message
  '''
  def setGraphics(self, graphics):
    if hasattr(graphics, "confirm"):
      if graphics.confirm("Extruder"):
        self.graphics = graphics
      else:
        return False, "%s:\ngraphics for ExtruderD is not extruderGraphics" % ERROR
    else:
      return False, "%s:\ninput is not a file for the ICSM program" % ERROR
    return True, ""

  '''
  The following function returns the socket that connects to the server
  '''
  def getSocket(self):
    return self.socket

  '''
  The following function returns the file name of the workflow sheet that is currently being selected within the 
  "Extruder" panel
  '''
  def getWorkflowFileName(self):
    return self.workflowFileName

  '''
  The following function sets the file name of the workflow sheet that is currently being selected within the 
  "Extruder" panel.
  '''
  def setWorkflowFileName(self, workflowFileName):
    self.workflowFileName = workflowFileName

  '''
  The following function returns the variable for the extruder drop down menu
  '''
  def getExtruderVariable(self):
    return self.extruderVariable

  '''
  The following function sets the variable for the extruder drop down menu
  '''
  def setExtruderVariable(self, extruderVariable):
    self.extruderVariable = extruderVariable

  '''
  The following function returns the variable for the die options drop down menu
  '''
  def getDieVariable(self):
    return self.dieVariable

  '''
  The following function sets the variable for the die options drop down menu
  '''
  def setDieVariable(self, dieVariable):
    self.dieVariable = dieVariable

  '''
  The following function returns the variable for the pre-die drop down menu
  '''
  def getPreDieVariable(self):
    return self.preDieVariable

  '''
  The following function sets the variable for the pre-die drop down menu
  '''
  def setPreDieVariable(self, preDieVariable):
    self.preDieVariable = preDieVariable

  '''
  The following function returns the variable for the screen pack check box
  '''
  def getScreenPackVariable(self):
    return self.screenVariable

  '''
  The following function sets the variable for the screen pack check box
  '''
  def setScreenPackVariable(self, screenVariable):
    self.screenVariable = screenVariable

  '''
  The following function returns the list of variables for the port drop down menus and their location on the extruder
  '''
  def getPortVariables(self):
    return self.portVariables

  '''
  The following function sets the list of variables for the port drop down menus and their location on the extruder
  '''
  def setPortVariables(self, portVariables):
    self.portVariables = portVariables

  '''
  The following function returns the list of variables for the side drop down menus and their location on the extruder
  '''
  def getSideVariables(self):
    return self.sideVariables

  '''
  The following function sets the list of variables for the side drop down menus and their location on the extruder
  '''
  def setSideVariables(self, sideVariables):
    self.sideVariables = sideVariables

  '''
  The following function returns the dictionary of feeder models for this instance
  '''
  def getFeeders(self):
    return self.feeders

  '''
  The following function returns the variable for the strand cooling frame
  '''
  def getStrandCoolingFrameVariable(self):
    return self.strandCoolingFrameVariable

  '''
  The following function sets the variable for the strand cooling frame
  '''
  def setStrandCoolingFrameVariable(self, strandCoolingFrameVariable):
    self.strandCoolingFrameVariable = strandCoolingFrameVariable

  '''
  The following function returns the variable for the strand separator
  '''
  def getSeparatorVariable(self):
    return self.separatorVariable

  '''
  The following function sets the variable for the strand separator
  '''
  def setSeparatorVariable(self, separatorVariable):
    self.separatorVariable = separatorVariable

  '''
  The following function returns the variable for the number of fans used on the extruder
  '''
  def getFansVariable(self):
    return self.fansVariable

  '''
  The following function sets the variable for the number of fans used on the extruder
  '''
  def setFansVariable(self, fansVariable):
    self.fansVariable = fansVariable

  '''
  The following function returns the variable that holds if the operator is using air knives on the extruder
  '''
  def getAirKnivesVariable(self):
    return self.airKnivesVariable

  '''
  The following function sets the variable that holds if the operator is using air knives on the extruder
  '''
  def setAirKnivesVariable(self, airKnivesVariable):
    self.airKnivesVariable = airKnivesVariable

  '''
  The following function returns the variable for the number of air knives used on the extruder
  '''
  def getNumAirKnivesVariable(self):
    return self.numAirKnivesVariable

  '''
  The following function sets the variable for the number of air knives used on the extruder
  '''
  def setNumAirKnivesVariable(self, numAirKnivesVariable):
    self.numAirKnivesVariable = numAirKnivesVariable

  '''
  The following function returns the list of belt mister varibles
  '''
  def getBeltMisters(self):
    return self.beltMisters

  '''
  The following function sets the list of belt mister varibles
  '''
  def setBeltMisters(self, beltMisters):
    self.beltMisters = beltMisters

  '''
  The following function returns the list of sluice mister varibles
  '''
  def getSluiceMisters(self):
    return self.sluiceMisters

  '''
  The following function sets the list of sluice mister varibles
  '''
  def setSluiceMisters(self, sluiceMisters):
    self.sluiceMisters = sluiceMisters

  '''
  The following function returns the variable for the pelletizer drop down menu
  '''
  def getPelletizierVariable(self):
    return self.pelletizierVariable

  '''
  The following function sets the variable for the pelletizer drop down menu
  '''
  def setPelletizierVariable(self, pelletizierVariable):
    self.pelletizierVariable = pelletizierVariable

  '''
  The following function returns the variable for the classified radio buttons
  '''
  def getClassifiedVariable(self):
    return self.classifiedVariable

  '''
  The following function sets the variable for the classified radio buttons
  '''
  def setClassifiedVariable(self, classifiedVariable):
    self.classifiedVariable = classifiedVariable

  '''
  The following function returns the variable for the data point every menu
  '''
  def getDataShiftVariable(self):
    return self.dataShiftVariable

  '''
  The following function sets the variable for the data point every menu
  '''
  def setDataShiftVariable(self, dataShiftVariable):
    self.dataShiftVariable = dataShiftVariable

  '''
  The following function returns the variable for the num of data points menu
  '''
  def getNumDataVariable(self):
    return self.numDataVariable

  '''
  The following function returns the variable for the num of data points menu
  '''
  def setNumDataVariable(self, numDataVariable):
    self.numDataVariable = numDataVariable

  '''
  The following function returns the variable for the from radio buttons
  '''
  def getFromVariable(self):
    return self.fromVariable

  '''
  The following function sets the variable for the from radio buttons
  '''
  def setFromVariable(self, fromVariable):
    self.fromVariable = fromVariable

  '''
  The following function returns the variable for the to radio buttons
  '''
  def getToVariable(self):
    return self.toVariable

  '''
  The following function sets the variable for the to radio buttons
  '''
  def setToVariable(self, toVariable):
    self.toVariable = toVariable

  '''
  The following function builds and returns a feeder model
  '''
  def buildFeeder(self):
    return feederData.FeederD(self.getConfig(), self.getTotalPercentage())

  '''
  The following function returns a boolen value dictating on whether or not the application is connected to the server
  '''
  def areConnected(self):
    try:
      self.getSocket().send("check")
      x = self.getSocket().recv(1024)
      return True
    except:
      return False

  '''
  The following function returns a boolean value dictating on whether or not the server is mounted to the local machine
  '''
  def areMounted(self):
    if os.path.isdir("/Volumes/Workflow"):
      return True
    return False

  '''
  The following function returns the total percentage of the materials in the feeder
  '''
  def getTotalPercentage(self):
    total = 0.0
    for frame, feeder in self.getFeeders().iteritems():
      total = total + feeder["Total"]
    return total

  '''
  The following function
  '''
  def listFeeders(self):

    # Get feeders from data
    feeders = self.getFeeders()
    list = []

    # Constuct the list by filling it with the values of the dictionary
    for key, value in feeders.iteritems():
      list.append(value)

    # Return the list of feeders
    return list

  '''
  The following function turns a list of variables into a list of values
  '''
  def __toValues(self, input):
    list = []
    for index in input:
      list.append([index[0], index[1].get()])
    return list

  '''
  The following function checks the data to see if the data is ready to update the workflow .xlsx sheet
  '''
  def checkForUpdate(self):

    # Build the initial error variables and values
    ready = True
    errors = []
    errorTextLabels = {}

    # Check the Extruder Section
    errorTextLabels[self.getConfig().EXTRUDER_OPTIONS_SECTION_TITLE] = []
    labels = errorTextLabels[self.getConfig().EXTRUDER_OPTIONS_SECTION_TITLE]

    # Check the Extruder drop down menu
    if self.getExtruderVariable().get() == self.getConfig().EXTRUDERS[0]:
      ready = False
      errors.append("Choose an \"Extruder\"\n")
      labels.append("Extruder")
    else:
      if self.getDieVariable().get() == self.getConfig().EXTRUDER_DIES[self.getExtruderVariable().get()][0]:
        ready = False
        errors.append("Choose a \"Die Option\"\n")
        labels.append("Die Options")
    if self.getPreDieVariable().get() == self.getConfig().PRE_DIE[0]:
      ready = False
      errors.append("Choose an \"Pre-Die Option\"\n")
      labels.append("Pre-Die")
    if not self.getGraphics().getMeshEntry().get():
      ready = False
      errors.append("Enter message into \"Mesh\"\n")
      labels.append("Mesh:")

    # Check the feeders section
    errorTextLabels[self.getConfig().FEEDERS_SECTION_TITLE] = []
    labels = errorTextLabels[self.getConfig().FEEDERS_SECTION_TITLE]

    # Check to make sure that there is at least one feeder in the list
    if len(self.getFeeders()) is 0:
      ready = False
      errors.append("Need at least one \"Feeder\"\n")
      labels.append("")

    # Check the strand cooling section
    errorTextLabels[self.getConfig().STRAND_COOLING_OPTIONS_SECTION_TITLE] = []
    labels = errorTextLabels[self.getConfig().STRAND_COOLING_OPTIONS_SECTION_TITLE]

    # Check the strand cooling radio buttons on the top of the section frame
    if self.getStrandCoolingFrameVariable().get() is 0:
      ready = False
      errors.append("Choose a \"Strand Cooling\" method\n")
      labels.append("")
    elif self.getStrandCoolingFrameVariable().get() is 2:
      if self.getAirKnivesVariable().get():
        if self.getGraphics().getLocationEntry().get() == "":
          ready = False
          errors.append("Enter message into Air Knives: \"Location\"\n")
          labels.append("Location:")
    elif self.getStrandCoolingFrameVariable().get() is 3:
      if self.getGraphics().getLengthEntry().get() == "":
        ready = False
        errors.append("Enter message into \"Length of Dip\"\n")
        labels.append("Length of Dip:")
      if self.getAirKnivesVariable().get():
        if self.getGraphics().getLocationEntry().get() == "":
          ready = False
          errors.append("Enter message into Air Knives: \"Location\"\n")
          labels.append("Location:")

    # Check the pelletizing section
    errorTextLabels[self.getConfig().PELLETIZING_OPTIONS_SECTION_TITLE] = []
    labels = errorTextLabels[self.getConfig().PELLETIZING_OPTIONS_SECTION_TITLE]

    # Check the pelletizier variable
    if self.getPelletizierVariable().get() == self.getConfig().PELLETIZIERS[0]:
      ready = False
      errors.append("Choose a \"Pelletizier\"\n")
      labels.append("Pelletizier")

    # Check the classified section
    errorTextLabels[self.getConfig().CLASSIFIED_OPTIONS_SECTION_TITLE] = []
    labels = errorTextLabels[self.getConfig().CLASSIFIED_OPTIONS_SECTION_TITLE]

    # Check the classified variable
    if self.getClassifiedVariable().get() is 0:
      ready = False
      errors.append("Choose a \"Classified\" option\n")
      labels.append("Classified:")

    # Return the error variables
    return ready, errors, errorTextLabels

  '''
  The following function checks to make sure that the program is ready to capture data from the extruder
  '''
  def checkForCapture(self):

    # Build the initial error variables and values
    ready = True
    errors = []
    errorTextLabels = {}

    # Add the section label
    errorTextLabels[self.getConfig().CAPTURING_SECTION_TITLE] = []
    labels = errorTextLabels[self.getConfig().CAPTURING_SECTION_TITLE]

    # Check capture data point menu variables
    if self.getDataShiftVariable().get() == self.getConfig().DATA_POINT_EVERY[0]:
      ready = False
      errors.append("Choose how often you want a \"Data Point\"\n")
      labels.append(self.getConfig().DATA_POINT_EVERY_LABEL)
    if self.getNumDataVariable().get() == self.getConfig().NUM_OF_DATA_POINTS[0]:
      ready = False
      errors.append("Choose total # of \"Data Points\"\n")
      labels.append(self.getConfig().NUM_OF_DATA_POINTS_LABEL)

    # Return the error variables
    return ready, errors, errorTextLabels

  '''
  The following function updates the workflow .xlsx sheets
  '''
  def updateWorkflow(self):

    # Build the initial dictionary variable
    update = {}

    # Add the data on the Extruder Section
    section = self.getConfig().EXTRUDER_OPTIONS_SECTION_TITLE
    update[section] = {"Extruder": self.getExtruderVariable().get()}
    update[section]["Die"] = self.getDieVariable().get()
    update[section]["Pre-Die"] = self.getPreDieVariable().get()
    update[section]["Screen Pack"] = self.getScreenPackVariable().get()
    update[section]["Mesh"] = self.getGraphics().getMeshEntry().get()

    # Add the data on the Port Set-Up Section
    section = self.getConfig().PORT_OPTIONS_SECTION_TITLE
    update[section] = {"Ports": self.__toValues(self.getPortVariables())}
    update[section]["Sides"] = self.__toValues(self.getSideVariables())

    # Add the data on the Feeders
    section = self.getConfig().FEEDERS_SECTION_TITLE
    update[section] = {"Feeders": self.listFeeders()}

    # Add the data on the Strand Cooling Section
    section = self.getConfig().STRAND_COOLING_OPTIONS_SECTION_TITLE
    update[section] = {"Strand Cooling": self.getStrandCoolingFrameVariable().get()}
    if self.getStrandCoolingFrameVariable().get() is 1 or self.getStrandCoolingFrameVariable().get() is 2 or\
         self.getStrandCoolingFrameVariable().get() is 3:
      update[section]["Separator"] = self.getSeparatorVariable().get()
      update[section]["Fans"] = self.getFansVariable().get()
    if self.getStrandCoolingFrameVariable().get() is 2 or self.getStrandCoolingFrameVariable().get() is 3:
      update[section]["AirKnives"] = self.getAirKnivesVariable().get()
      if self.getAirKnivesVariable().get():
        update[section]["Location"] = self.getGraphics().getLocationEntry().get()
        update[section]["AirKnivesNum"] = self.getNumAirKnivesVariable().get()
    if self.getStrandCoolingFrameVariable().get() is 2 or self.getStrandCoolingFrameVariable().get() is 4:
      update[section]["WaterTemp"] = self.getGraphics().getWaterTempScale().get()
    if self.getStrandCoolingFrameVariable().get() is 3:
      update[section]["LengthOfBath"] = self.getGraphics().getLengthEntry().get()
    if self.getStrandCoolingFrameVariable().get() is 4:
      update[section]["BeltMisters"] = self.__toValues(self.getBeltMisters())
      update[section]["SluiceMisters"] = self.__toValues(self.getSluiceMisters())
      update[section]["Conveyor"] = self.getGraphics().getConveyorScale().get()
      update[section]["Blower"] = self.getGraphics().getBlowerScale().get()
      update[section]["BlowerVac"] = self.getGraphics().getVacBlowerScale().get()

    # Add the data on the Pelletizing Section
    section = self.getConfig().PELLETIZING_OPTIONS_SECTION_TITLE
    update[section] = {"Pelletizier": self.getPelletizierVariable().get()}
    update[section]["Feed Roll"] = self.getGraphics().getFeedRollScale().get()
    update[section]["Rotor"] = self.getGraphics().getRotorScale().get()
    if self.getPelletizierVariable().get() == self.getConfig().PELLETIZIERS[2]:
      update[section]["Feeder Speed"] = self.getGraphics().getFeederSpeedScale().get()
      update[section]["Pump Speed"] = self.getGraphics().getPumpSpeedScale().get()
    update[section]["Comments"] = self.getGraphics().getPelletizingComments().get(1.0, "end-1c")

    # Add the data on the Classified Section
    section = self.getConfig().CLASSIFIED_OPTIONS_SECTION_TITLE
    names = self.getConfig().CLASSIFIED_RADIO_BUTTON_NAMES
    update[section] = {"Classified": names[self.getClassifiedVariable().get() - 1]}

    # Check workflow file name
    if self.getWorkflowFileName() is None:
      self.getGraphics().displayError("Please Select a Workflow")
      return 0

    # Check on Server connection
    if self.areConnected():
      # FINISH CODE
      if self.editXLSXFile(update):
        self.getGraphics().displayMessage("Workflow was Updated")
      else:
        self.getGraphics().displayError("Workflow was not updated")
    elif self.areMounted():
      if self.editXLSXFile(update):
        self.getGraphics().displayMessage("Workflow was Updated")
      else:
        self.getGraphics().displayError("Workflow was not updated")
    else:
      self.getGraphics().displayError("The Appication is Unable to Connect to the Workflow")
      return 0

  '''
  The following function tells the program on the server to capture data from the extruder and add it to the workflow
  '''
  def captureExtruder(self):

    # Build the initial dictionary variable
    capture = {}

    # Check workflow file name
    if self.getWorkflowFileName() is None:
      self.getGraphics().displayError("Please Select a Workflow")
      return 0

    # Add the data on the Capturing Section
    section = self.getConfig().CAPTURING_SECTION_TITLE
    capture[section] = {"Data": self.getDataShiftVariable().get()}
    capture[section]["Num"] = self.getNumDataVariable().get()
    if self.getGraphics().getFromEntry().get():
      capture[section]["From"] = [self.getGraphics().getFromEntry().get(), self.getFromVariable().get()]
    else:
      capture[section]["From"] = 0
    if self.getGraphics().getToEntry().get():
      capture[section]["To"] = [self.getGraphics().getToEntry().get(), self.getToVariable().get()]
    else:
      capture[section]["To"] = 0
    capture[section]["Date"] = self.getGraphics().getCalendarEntry().get()

    # Check on Server connection
    if self.areConnected():
      self.getSocket().send("capture")
      print self.getSocket().recv(1024)
      self.getSocket().send("27Entek")
      print self.getSocket().recv(1024)
      self.getSocket().send(self.getWorkflowFileName())
      print self.getSocket().recv(1024)
      self.getSocket().send(json.dumps(capture))
      print "Received Message: ", self.getSocket().recv(1024)
    else:
      self.getGraphics().displayError("The Appication is Unable to Connect to the Server")
      return 0

  '''
  The following function updates the workflow with the edited information
  '''
  def editXLSXFile(self, update):

    # Open the Workflow
    wb = openpyxl.load_workbook("/Volumes/Workflow/TK_WF_DONT_TOUCH/wf_template_new.xlsx")

    # Edit the extruding area
    extruderDict = update["Extruder Options"]
    wb["TK_Compounding"]["A6"] = extruderDict["Extruder"]
    wb["TK_Compounding"]["C6"] = extruderDict["Die"].split("-")[0]
    wb["TK_Compounding"]["D6"] = extruderDict["Die"].split("-")[1]
    wb["TK_Compounding"]["F6"] = extruderDict["Pre-Die"]
    if extruderDict["Screen Pack"] is 1:
      wb["TK_Compounding"]["D8"] = "Yes"
    else:
      wb["TK_Compounding"]["D8"] = "No"
    wb["TK_Compounding"]["D9"] = extruderDict["Mesh"]

    # Edit the Port set-up area
    portDict = update["Port Set-Up"]
    ports = portDict["Ports"]
    sides = portDict["Sides"]

    # Edit the Feeder area
    feederDict = update["Feeders"]
    feeders = feederDict["Feeders"]
    for row in [13, 14, 15, 20, 25, 30, 35, 37, 62, 70]:
      for column in range(1, 22):
        wb["TK_Compounding"].cell(row=row, column=column).border = Border(bottom=Side(border_style="medium",
                                                                                      color=colors.BLACK))
    index = 0
    for feeder in feeders:
      row = (index * 5) + 16
      wb["TK_Compounding"]["B" + str(row)] = feeder["Feeder"]
      if "Tube" in feeder:
        wb["TK_Compounding"]["E" + str(row)] = feeder["Tube"]
      wb["TK_Compounding"]["G" + str(row)] = feeder["Screw"]
      if feeder["Location"] is 2:
        wb["TK_Compounding"]["L" + str(row)] = "Side Stuffer"
      else:
        wb["TK_Compounding"]["L" + str(row)] = "Throat"
      index2 = 0
      for RM, per in feeder["RM"].iteritems():
        wb["TK_Compounding"]["N" + str(row + index2)] = RM
        wb["TK_Compounding"]["R" + str(row + index2)] = str(per)
        if feeder["Set Point"] is 1:
          wb["TK_Compounding"]["T" + str(row + index2)] = "lbs/hr"
        else:
          wb["TK_Compounding"]["T" + str(row + index2)] = "kg/hr"
        index2 = index2 + 1
      index = index + 1

    # Edit the Strand Cooling area
    strandDict = update["Strand Cooling Options"]
    if strandDict["Strand Cooling"] is 1:
      cooling = "Belt"
    elif strandDict["Strand Cooling"] is 2:
      cooling = "Belt with Mister"
    elif strandDict["Strand Cooling"] is 3:
      cooling = "Water Bath"
    elif strandDict["Strand Cooling"] is 4:
      cooling = "Spray Belt"
    elif strandDict["Strand Cooling"] is 5:
      cooling = "UWP"
    else:
      cooling = "Other"
    wb["TK_Compounding"]["L6"] = cooling
    wb["TK_Compounding"]["E64"] = cooling
    if strandDict["Strand Cooling"] is 1 or strandDict["Strand Cooling"] is 2 or strandDict["Strand Cooling"] is 3:
      if strandDict["Separator"] is 1:
        wb["TK_Compounding"]["E66"] = "Yes"
      else:
        wb["TK_Compounding"]["E66"] = "No"
      wb["TK_Compounding"]["I66"] = strandDict["Fans"]
    if strandDict["Strand Cooling"] is 2 or strandDict["Strand Cooling"] is 3:
      if strandDict["AirKnives"] is 1:
        wb["TK_Compounding"]["S64"] = "Yes"
        wb["TK_Compounding"]["S65"] = strandDict["AirKnivesNum"]
        wb["TK_Compounding"]["R66"] = strandDict["Location"]
      else:
        wb["TK_Compounding"]["S64"] = "No"
    if strandDict["Strand Cooling"] is 2:
      wb["TK_Compounding"]["K66"] = "Water Temp:"
      wb["TK_Compounding"]["M66"] = strandDict["WaterTemp"]
    if strandDict["Strand Cooling"] is 3:
      wb["TK_Compounding"]["K66"] = "Length of Dip"
      wb["TK_Compounding"]["M66"] = strandDict["LengthOfBath"]
    if strandDict["Strand Cooling"] is 4:
      wb["TK_Compounding"]["R69"] = strandDict["Conveyor"]
      wb["TK_Compounding"]["S69"] = strandDict["Blower"]
      wb["TK_Compounding"]["T69"] = strandDict["BlowerVac"]
      wb["TK_Compounding"]["U69"] = strandDict["WaterTemp"]
      index = 0
      for mister in strandDict["BeltMisters"]:
        if mister[1] is 1:
          row = 69
          value = "On"
        else:
          row = 70
          value = "Off"
        wb["TK_Compounding"].cell(row=row, column=(3 + index), value=value)
        index = index + 1
      index = 0
      for sluice in strandDict["SluiceMisters"]:
        if sluice[1] is 1:
          row = 69
          value = "On"
        else:
          row = 70
          value = "Off"
        wb["TK_Compounding"].cell(row=row, column=(16 + index), value=value)
        index = index + 1

    # Edit the Pelletizier area
    pelletizierDict = update["Pelletizier Options"]
    wb["TK_Compounding"]["P6"] = pelletizierDict["Pelletizier"]
    wb["TK_Compounding"]["S6"] = pelletizierDict["Feed Roll"]
    wb["TK_Compounding"]["S8"] = pelletizierDict["Rotor"]
    if "Feeder Speed" in pelletizierDict:
      feederSpeed = pelletizierDict["Feeder Speed"]
    if "Pump Speed" in pelletizierDict:
      pumpSpeed = pelletizierDict["Pump Speed"]
    comments = pelletizierDict["Comments"]

    # Edit the classified area
    classifiedDict = update["Classified Options"]
    wb["TK_Compounding"]["U6"] = classifiedDict["Classified"]

    # Save the Workflow
    wb.save("/Volumes/Workflow/TK_WF_DONT_TOUCH/wf_modified.xlsx")

    return True

  '''
  The following function returns a confirmation that tells the calling code which class file this function belongs to
  '''
  def confirm(self, value):
    if isinstance(value, basestring):
      if value.lower() == "extruder":
        return True
      else:
        return False
    else:
      return False